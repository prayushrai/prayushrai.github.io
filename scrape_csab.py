#!/usr/bin/env python3
"""
CSAB Special Round 2025 cutoff scraper.
Source: https://admissions.nic.in/csabspl/Applicant/seatallotmentresult/currentorcr.aspx

Scrapes the FINAL round (round 3) for NIT, IIIT (3IT), GFTI (CFI) and writes
per-type CSVs, a combined CSV, and a styled XLSX.
"""
import os
import re
import time
import requests
import pandas as pd
from bs4 import BeautifulSoup

URL = "https://admissions.nic.in/csabspl/Applicant/seatallotmentresult/currentorcr.aspx"
PFX = "ctl00$ContentPlaceHolder1$"
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
    ),
    "Referer": URL,
    "Origin": "https://admissions.nic.in",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}

OUT_DIR = "/sessions/amazing-cool-knuth/mnt/outputs"

ROUND = "3"  # final CSAB Special 2025 round

# Map: instype option-value (with trailing space as returned by site) -> short tag
INSTYPE_MAP = [
    ("NIT ", "nit"),
    ("3IT ", "iiit"),
    ("CFI ", "gfti"),
]


def parse_state(html):
    soup = BeautifulSoup(html, "lxml")
    state = {}
    for f in ("__VIEWSTATE", "__VIEWSTATEGENERATOR", "__EVENTVALIDATION"):
        el = soup.find("input", {"name": f})
        state[f] = el["value"] if el and el.has_attr("value") else ""
    return state, soup


def post(session, data, sleep=1.0):
    time.sleep(sleep)
    r = session.post(URL, data=data, timeout=60)
    r.raise_for_status()
    return r


def fetch_grid_for_instype(session, instype_value):
    """Walk the dropdown postback chain and return the GridView table soup."""
    r = session.get(URL, timeout=60)
    r.raise_for_status()
    state, _ = parse_state(r.text)

    # round
    data = {
        "__EVENTTARGET": PFX + "ddlroundno",
        "__EVENTARGUMENT": "",
        "__LASTFOCUS": "",
        **state,
        PFX + "ddlroundno": ROUND,
    }
    r = post(session, data); state, _ = parse_state(r.text)

    # instype
    data = {
        "__EVENTTARGET": PFX + "ddlInstype",
        "__EVENTARGUMENT": "",
        "__LASTFOCUS": "",
        **state,
        PFX + "ddlroundno": ROUND,
        PFX + "ddlInstype": instype_value,
    }
    r = post(session, data); state, _ = parse_state(r.text)

    # institute = ALL
    data = {
        "__EVENTTARGET": PFX + "ddlInstitute",
        "__EVENTARGUMENT": "",
        "__LASTFOCUS": "",
        **state,
        PFX + "ddlroundno": ROUND,
        PFX + "ddlInstype": instype_value,
        PFX + "ddlInstitute": "ALL",
    }
    r = post(session, data); state, _ = parse_state(r.text)

    # branch = ALL (Seattype dropdown stays empty in CSAB)
    data = {
        "__EVENTTARGET": PFX + "ddlBranch",
        "__EVENTARGUMENT": "",
        "__LASTFOCUS": "",
        **state,
        PFX + "ddlroundno": ROUND,
        PFX + "ddlInstype": instype_value,
        PFX + "ddlInstitute": "ALL",
        PFX + "ddlBranch": "ALL",
    }
    r = post(session, data); state, _ = parse_state(r.text)

    # submit with Seattype=ALL
    data = {
        "__EVENTTARGET": "",
        "__EVENTARGUMENT": "",
        "__LASTFOCUS": "",
        **state,
        PFX + "ddlroundno": ROUND,
        PFX + "ddlInstype": instype_value,
        PFX + "ddlInstitute": "ALL",
        PFX + "ddlBranch": "ALL",
        PFX + "ddlSeattype": "ALL",
        PFX + "btnSubmit": "Submit",
    }
    r = post(session, data)
    soup = BeautifulSoup(r.text, "lxml")
    tbl = (
        soup.find("table", id=re.compile("GridView", re.I))
        or soup.find("table", id=re.compile("gv", re.I))
        or soup.select_one("#ctl00_ContentPlaceHolder1_GridView1")
    )
    return tbl


def parse_grid(tbl):
    if tbl is None:
        return pd.DataFrame()
    rows = tbl.find_all("tr")
    if len(rows) < 2:
        return pd.DataFrame()
    headers = [c.get_text(strip=True) for c in rows[0].find_all(["td", "th"])]
    data = []
    for tr in rows[1:]:
        cells = [c.get_text(strip=True) for c in tr.find_all(["td", "th"])]
        if len(cells) == len(headers):
            data.append(cells)
    return pd.DataFrame(data, columns=headers)


def to_int_rank(v):
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    s = re.sub(r"P$", "", s, flags=re.I).strip()
    try:
        return int(s)
    except ValueError:
        try:
            return int(re.sub(r"[^\d]", "", s))
        except ValueError:
            return None


def add_int_columns(df):
    if df.empty:
        return df
    if "Opening Rank" in df.columns:
        df["Opening Rank (int)"] = df["Opening Rank"].map(to_int_rank)
    if "Closing Rank" in df.columns:
        df["Closing Rank (int)"] = df["Closing Rank"].map(to_int_rank)
    return df


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    session = requests.Session()
    session.headers.update(HEADERS)

    frames = {}
    for instype_value, tag in INSTYPE_MAP:
        print(f"[+] Fetching {tag} (instype={instype_value!r})...", flush=True)
        tries = 0
        df = pd.DataFrame()
        while tries < 3:
            try:
                tbl = fetch_grid_for_instype(session, instype_value)
                df = parse_grid(tbl)
                if not df.empty:
                    break
                print(f"    empty result, retry {tries+1}", flush=True)
            except Exception as e:
                print(f"    error: {e}; retry {tries+1}", flush=True)
            tries += 1
            time.sleep(2)
        df = add_int_columns(df)
        out_csv = os.path.join(OUT_DIR, f"csab_2025_final_{tag}.csv")
        df.to_csv(out_csv, index=False)
        print(f"    -> {out_csv}  rows={len(df)}", flush=True)
        frames[tag] = df

    label_map = {"nit": "NIT", "iiit": "IIIT", "gfti": "GFTI"}
    parts = []
    for tag, df in frames.items():
        if df.empty:
            continue
        d = df.copy()
        d.insert(0, "Institute Type", label_map[tag])
        parts.append(d)
    combined = pd.concat(parts, ignore_index=True) if parts else pd.DataFrame()

    all_csv = os.path.join(OUT_DIR, "csab_2025_final_all.csv")
    combined.to_csv(all_csv, index=False)
    print(f"[+] Combined CSV -> {all_csv}  rows={len(combined)}", flush=True)

    all_xlsx = os.path.join(OUT_DIR, "csab_2025_final_all.xlsx")
    with pd.ExcelWriter(all_xlsx, engine="openpyxl") as writer:
        combined.to_excel(writer, sheet_name="CSAB 2025 Final", index=False)
        ws = writer.sheets["CSAB 2025 Final"]
        ws.freeze_panes = "A2"
        if len(combined) > 0:
            from openpyxl.utils import get_column_letter
            last_col = get_column_letter(len(combined.columns))
            ws.auto_filter.ref = f"A1:{last_col}{len(combined)+1}"
            for i, col in enumerate(combined.columns, start=1):
                try:
                    maxlen = int(combined[col].astype(str).str.len().clip(upper=60).max())
                except Exception:
                    maxlen = 12
                ws.column_dimensions[get_column_letter(i)].width = min(60, max(12, maxlen + 2))
    print(f"[+] XLSX -> {all_xlsx}", flush=True)

    print("\n=== SUMMARY ===")
    print(f"Round used: {ROUND}")
    for tag, df in frames.items():
        print(f"  {label_map[tag]:5s}: {len(df)} rows")
    print(f"  TOTAL: {len(combined)} rows")
    if not combined.empty:
        print("First row:", combined.iloc[0].to_dict())


if __name__ == "__main__":
    main()
